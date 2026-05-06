"""Exporter tests: refuse-to-emit + artifact production."""

from __future__ import annotations

import zipfile
from decimal import Decimal
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from accounting_parser.exporters import (
    CCHEngagementExporter,
    RefuseToEmit,
)
from accounting_parser.model.canonical import (
    Account,
    AccountType,
    WorkingTrialBalance,
    WTBRow,
)


def _wtb_all_classified() -> WorkingTrialBalance:
    return WorkingTrialBalance(engagement_id=uuid4(), rows=(
        WTBRow(
            account=Account(
                account_number="1000", account_name="Cash",
                account_type=AccountType.ASSET, category="Cash",
            ),
            unadjusted=Decimal("1000.00"),
            adjusted=Decimal("1000.00"),
            final=Decimal("1000.00"),
            tax_basis=Decimal("1000.00"),
        ),
        WTBRow(
            account=Account(
                account_number="4000", account_name="Revenue",
                account_type=AccountType.REVENUE, category="Revenue",
            ),
            unadjusted=Decimal("5000.00"),
            adjusted=Decimal("5000.00"),
            final=Decimal("5000.00"),
            tax_basis=Decimal("5000.00"),
        ),
    ))


def _wtb_with_unclassified() -> WorkingTrialBalance:
    return WorkingTrialBalance(engagement_id=uuid4(), rows=(
        WTBRow(
            account=Account(
                account_number="9999", account_name="Suspense",
                account_type=None, category=None,
            ),
            unadjusted=Decimal("100"),
            adjusted=Decimal("100"),
            final=Decimal("100"),
            tax_basis=Decimal("100"),
        ),
    ))


def test_cch_export_refuses_with_unclassified() -> None:
    exporter = CCHEngagementExporter()
    wtb = _wtb_with_unclassified()
    blockers = exporter.validate(wtb)
    assert blockers, "expected blockers for unclassified account"
    assert any("Unclassified" in b.message or "unclassified" in b.message
               for b in blockers)


def test_cch_emit_refuses_on_blockers(tmp_path: Path) -> None:
    exporter = CCHEngagementExporter()
    wtb = _wtb_with_unclassified()
    with pytest.raises(RefuseToEmit):
        exporter.emit(wtb, tmp_path)


def test_cch_emit_produces_zip_with_three_artifacts(tmp_path: Path) -> None:
    exporter = CCHEngagementExporter(client_name="Acme", engagement_year=2025)
    wtb = _wtb_all_classified()
    result = exporter.emit(wtb, tmp_path)
    assert result.emitted
    assert len(result.artifacts) == 1
    zpath = result.artifacts[0]
    assert zpath.exists()
    with zipfile.ZipFile(zpath) as zf:
        names = set(zf.namelist())
    assert {"tb_import.xlsx", "TBLinkTrigger.dl", "xtblink_companion.xlsx"} <= names


def test_cch_tb_xlsx_has_correct_columns(tmp_path: Path) -> None:
    exporter = CCHEngagementExporter()
    wtb = _wtb_all_classified()
    exporter.emit(wtb, tmp_path)
    wb = load_workbook(tmp_path / "tb_import.xlsx", data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers[0] == "Account Number"
    assert headers[10] == "Tax Basis"
    assert len(headers) == 13
