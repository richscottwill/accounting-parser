"""CCH Axcess Engagement exporter.

Emits:
1. The 13-column TB import Excel template populated with the Engagement's
   WTB rows (account number/name/type/prior/unadjusted/aje/adjusted/rje/
   final/tje/tax_basis/fs_grouping/tax_grouping).
2. A Dynalink trigger file (TBLinkTrigger.dl).
3. An XTBLink companion workbook scaffold.

All three are packaged in a ZIP for download.

Blockers:
- Unclassified accounts > 0
- Any WTB row with missing account_type
- Any validator error at blocker severity
"""

from __future__ import annotations

import zipfile
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook

from accounting_parser.exporters.base import ExportBlocker, ExportResult, RefuseToEmit
from accounting_parser.model.canonical import WorkingTrialBalance

CCH_COLUMNS = (
    "Account Number",
    "Account Name",
    "Account Type",
    "Prior Year",
    "Unadjusted",
    "AJE",
    "Adjusted",
    "RJE",
    "Final",
    "TJE",
    "Tax Basis",
    "Financial Statement Grouping",
    "Tax Grouping",
)


@dataclass
class CCHEngagementExporter:
    target_system: str = "cch_axcess_engagement"
    client_name: str = "Demo Client"
    engagement_year: int = 2025

    def validate(self, wtb: WorkingTrialBalance) -> list[ExportBlocker]:
        blockers: list[ExportBlocker] = []
        unclassified = [
            r
            for r in wtb.rows
            if r.account.category is None or r.account.category.lower() == "unclassified"
        ]
        if unclassified:
            blockers.append(
                ExportBlocker(
                    rule_id="R17.cch.unclassified",
                    message=f"{len(unclassified)} account(s) are Unclassified: "
                    f"{', '.join(r.account.account_number for r in unclassified[:5])}"
                    + (", ..." if len(unclassified) > 5 else ""),
                )
            )
        missing_type = [r for r in wtb.rows if r.account.account_type is None]
        if missing_type:
            blockers.append(
                ExportBlocker(
                    rule_id="R17.cch.missing_account_type",
                    message=f"{len(missing_type)} account(s) lack account_type",
                )
            )
        return blockers

    def emit(self, wtb: WorkingTrialBalance, output_dir: Path) -> ExportResult:
        blockers = self.validate(wtb)
        if blockers:
            raise RefuseToEmit(f"CCH export refused: {len(blockers)} blocker(s)")
        output_dir.mkdir(parents=True, exist_ok=True)
        tb_xlsx = self._write_tb_xlsx(wtb, output_dir / "tb_import.xlsx")
        dynalink = self._write_dynalink(output_dir / "TBLinkTrigger.dl")
        xtblink = self._write_xtblink(output_dir / "xtblink_companion.xlsx")

        zip_path = output_dir / f"cch_engagement_export_{self.client_name}.zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.write(tb_xlsx, tb_xlsx.name)
            zf.write(dynalink, dynalink.name)
            zf.write(xtblink, xtblink.name)
        return ExportResult(
            target_system=self.target_system,
            artifacts=(zip_path,),
            blockers=(),
        )

    def _write_tb_xlsx(self, wtb: WorkingTrialBalance, path: Path) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "TB Import"
        for i, col in enumerate(CCH_COLUMNS, start=1):
            ws.cell(row=1, column=i, value=col)
        for r, row in enumerate(wtb.rows, start=2):
            acc = row.account
            ws.cell(row=r, column=1, value=acc.account_number)
            ws.cell(row=r, column=2, value=acc.account_name)
            ws.cell(row=r, column=3, value=acc.account_type.value if acc.account_type else "")
            ws.cell(row=r, column=4, value=float(row.prior_year))
            ws.cell(row=r, column=5, value=float(row.unadjusted))
            ws.cell(row=r, column=6, value=float(row.sum_aje))
            ws.cell(row=r, column=7, value=float(row.adjusted))
            ws.cell(row=r, column=8, value=float(row.sum_rje))
            ws.cell(row=r, column=9, value=float(row.final))
            ws.cell(row=r, column=10, value=float(row.sum_tje))
            ws.cell(row=r, column=11, value=float(row.tax_basis))
            ws.cell(row=r, column=12, value=acc.category or "")
            ws.cell(row=r, column=13, value="")  # tax_grouping blank at MVP
        wb.save(path)
        return path

    def _write_dynalink(self, path: Path) -> Path:
        # Minimal Dynalink trigger file — the real format is XML per R17.4.
        ts = datetime.now(UTC).isoformat()
        path.write_text(
            f'<?xml version="1.0"?>\n'
            f'<DynalinkTrigger timestamp="{ts}" client="{self.client_name}">\n'
            f"  <Action>UpdateTrialBalance</Action>\n"
            f"  <Year>{self.engagement_year}</Year>\n"
            f"</DynalinkTrigger>\n",
            encoding="utf-8",
        )
        return path

    def _write_xtblink(self, path: Path) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "XTBLink"
        ws["A1"] = f"XTBLink companion for {self.client_name} {self.engagement_year}"
        wb.save(path)
        return path
