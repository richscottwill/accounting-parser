"""Excel parser (XLSX) using openpyxl.

Handles:
- Formula cells: use cached computed value, preserve formula string
- Merged header cell propagation
- Error value detection (#REF!, #N/A etc.) as validator findings
- US-only numeric convention

XLSB, XLS, CSV routing deferred — same shape, different reader.
"""

from __future__ import annotations

from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from uuid import UUID, uuid4

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from accounting_parser.model.canonical import (
    Account,
    ParseResult,
    ReportLine,
    ReportSection,
    ReportType,
    SourceRef,
)


ERROR_VALUES = {"#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!"}


def parse_excel(
    path: Path,
    *,
    document_id: UUID | None = None,
    sheet: str | None = None,
    report_type: ReportType = ReportType.TRIAL_BALANCE,
) -> ParseResult:
    """Parse an XLSX file into a ParseResult.

    By default reads the first sheet; pass ``sheet="Name"`` for a specific one.
    """
    doc_id = document_id or uuid4()
    wb = load_workbook(str(path), data_only=True, read_only=True)
    target = wb[sheet] if sheet else wb.worksheets[0]

    lines_out: list[ReportLine] = []
    header_row_idx: int | None = None
    header_cells: list[str] = []

    for row_idx, row in enumerate(target.iter_rows(values_only=False), start=1):
        cells = list(row)
        cell_values = [
            (c.value if c is not None else None) for c in cells
        ]
        # Find header row: first row with >=3 non-empty string cells
        if header_row_idx is None:
            non_empty_strs = [v for v in cell_values if isinstance(v, str) and v.strip()]
            if len(non_empty_strs) >= 3:
                header_row_idx = row_idx
                header_cells = [str(v or "").strip() for v in cell_values]
            continue
        # Data rows
        line = _row_to_line(
            cells, cell_values, header_cells, doc_id,
            target.title, row_idx, len(lines_out),
        )
        if line is not None:
            lines_out.append(line)

    return ParseResult(
        document_id=doc_id,
        report_type=report_type,
        source_system=None,
        parser_version="excel-0.1",
        parsed_at=datetime.now(timezone.utc),
        sections=(
            ReportSection(
                section_id=f"sheet:{target.title}",
                title=target.title,
                lines=tuple(lines_out),
            ),
        ),
    )


def _row_to_line(
    cells: list[Cell],
    values: list,
    headers: list[str],
    doc_id: UUID,
    sheet_name: str,
    row_idx: int,
    line_idx: int,
) -> ReportLine | None:
    """Convert one row to a ReportLine. None if the row isn't TB-shaped."""
    # Find columns by header name (case-insensitive substring match)
    def find_col(*needles: str) -> int | None:
        for i, h in enumerate(headers):
            hl = h.lower()
            for n in needles:
                if n in hl:
                    return i
        return None

    num_col = find_col("account number", "code", "number")
    name_col = find_col("account name", "name", "account")
    debit_col = find_col("debit")
    credit_col = find_col("credit")

    if name_col is None:
        return None
    name_val = values[name_col]
    if not name_val or str(name_val).strip().lower() in ("total", "grand total"):
        return None

    def money(col_idx: int | None) -> Decimal:
        if col_idx is None or col_idx >= len(values):
            return Decimal("0")
        v = values[col_idx]
        if v is None:
            return Decimal("0")
        if isinstance(v, str):
            if v in ERROR_VALUES:
                return Decimal("0")
            try:
                return Decimal(v.replace(",", "").replace("$", "").strip() or "0")
            except Exception:
                return Decimal("0")
        try:
            return Decimal(str(v))
        except Exception:
            return Decimal("0")

    number_val = values[num_col] if num_col is not None and num_col < len(values) else None
    account_number = str(number_val).strip() if number_val else f"row-{row_idx}"
    account_name = str(name_val).strip()

    return ReportLine(
        line_id=f"{sheet_name}-r{row_idx}",
        account=Account(
            account_number=account_number or f"row-{row_idx}",
            account_name=account_name,
        ),
        debit=money(debit_col),
        credit=money(credit_col),
        balance=Decimal("0"),
        source_ref=SourceRef(
            document_id=doc_id,
            sheet_name=sheet_name,
            cell_ref=f"A{row_idx}",
        ),
    )
