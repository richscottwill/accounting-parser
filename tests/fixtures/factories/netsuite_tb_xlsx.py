"""NetSuite Trial Balance XLSX factory.

NetSuite exports are XLSX with multiple sheets; TB goes to a single sheet
with Subsidiary / Account / Debit / Credit columns and a Grand Total row.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from factories._data import DEFAULT_CHART, Account, balanced_debits_credits


def netsuite_tb_xlsx_factory(
    output_path: Path,
    *,
    company_name: str = "Synthetic Demo Parent, Inc.",
    subsidiary: str = "Synthetic Demo Co, LLC",
    period_end: str = "December 31, 2024",
    accounts: Sequence[Account] | None = None,
) -> Path:
    """Generate a NetSuite-style Trial Balance XLSX."""
    accs = tuple(accounts) if accounts is not None else DEFAULT_CHART
    debits, credits = balanced_debits_credits(accs)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "TrialBalance"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    ws["A1"] = company_name
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Subsidiary: {subsidiary}"
    ws["A3"] = f"Period: {period_end}"
    ws["A4"] = "Report: Trial Balance"

    headers = ["Subsidiary", "Account Number", "Account Name", "Type", "Debit", "Credit"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=6, column=i, value=h)
        c.font = bold
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    row = 7
    for a in accs:
        ws.cell(row=row, column=1, value=subsidiary)
        ws.cell(row=row, column=2, value=a.number)
        ws.cell(row=row, column=3, value=a.name)
        ws.cell(row=row, column=4, value=a.type)
        ws.cell(row=row, column=5, value=float(a.balance) if a.normal_balance == "debit" else None)
        ws.cell(row=row, column=6, value=float(a.balance) if a.normal_balance == "credit" else None)
        ws.cell(row=row, column=5).number_format = "#,##0.00;(#,##0.00)"
        ws.cell(row=row, column=6).number_format = "#,##0.00;(#,##0.00)"
        row += 1

    # Grand Total
    ws.cell(row=row, column=3, value="Grand Total").font = bold
    ws.cell(row=row, column=5, value=float(debits)).font = bold
    ws.cell(row=row, column=6, value=float(credits)).font = bold

    widths = {1: 26, 2: 14, 3: 34, 4: 12, 5: 16, 6: 16}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    wb.properties.creator = "NetSuite"
    wb.properties.title = f"Trial Balance - {subsidiary}"

    wb.save(output_path)
    return output_path
