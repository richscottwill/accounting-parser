"""Rejection-path fixture factories.

Produces inputs that the ingestion pipeline must hard-reject:
- Password-protected PDFs/XLSX
- Corrupted PDFs/XLSX (truncated, bad headers)
- Image-only scans (no text layer) — OCR path test fixture

These exercise the error-handling branches in Tasks 6 (ingestion) and 8-10
(parsing). Numeric values remain obvious-fake.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from pypdf import PdfReader, PdfWriter
from PIL import Image, ImageDraw
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas as rl_canvas


def _tmp_pdf(output_path: Path) -> Path:
    """Helper to make a throwaway small PDF."""
    c = rl_canvas.Canvas(str(output_path), pagesize=LETTER)
    c.setFont("Helvetica", 12)
    c.drawString(inch, 10 * inch, "Synthetic Source Document")
    c.drawString(inch, 9.5 * inch, "Cash - Operating          $123,456.78")
    c.drawString(inch, 9.0 * inch, "This is a rejection-path fixture.")
    c.showPage()
    c.save()
    return output_path


def password_protected_pdf_factory(
    output_path: Path, *, password: str = "SyntheticDemo2024"
) -> Path:
    """Generate a PDF encrypted with a user password."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    tmp = output_path.with_suffix(".unencrypted.pdf")
    _tmp_pdf(tmp)

    reader = PdfReader(str(tmp))
    writer = PdfWriter(clone_from=reader)
    writer.encrypt(user_password=password, owner_password=password, algorithm="AES-128")
    with output_path.open("wb") as f:
        writer.write(f)
    tmp.unlink()
    return output_path


def password_protected_xlsx_factory(
    output_path: Path, *, password: str = "SyntheticDemo2024"
) -> Path:
    """Generate a workbook-protected XLSX.

    openpyxl supports workbook structure protection; true at-rest XLSX
    encryption requires ``msoffcrypto-tool`` which isn't in our deps.
    Structure protection is enough to trigger the ingestion rejection
    path in Task 6 because the protection flag is detectable.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Protected"
    ws["A1"] = "Synthetic protected workbook"
    ws["A2"] = "This fixture is used to exercise the rejection path."
    wb.security.workbookPassword = password
    wb.security.lockStructure = True
    wb.save(output_path)
    return output_path


def corrupted_pdf_factory(output_path: Path) -> Path:
    """Generate a truncated (corrupted) PDF file.

    Writes the first ~400 bytes of a valid PDF then stops — no xref, no
    trailer. ``pdfplumber`` raises on open.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    tmp = output_path.with_suffix(".fullpdf.pdf")
    _tmp_pdf(tmp)
    data = tmp.read_bytes()
    tmp.unlink()

    # Take just the first 400 bytes: has %PDF header but no xref/trailer
    truncated = data[:400]
    output_path.write_bytes(truncated)
    return output_path


def corrupted_xlsx_factory(output_path: Path) -> Path:
    """Generate a corrupted XLSX (not a valid ZIP)."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    # XLSX is a ZIP file; writing non-ZIP bytes with an .xlsx extension
    # produces a file openpyxl will reject.
    output_path.write_bytes(b"PK\x03\x04NOT_A_VALID_XLSX_FILE_CORRUPTED_FIXTURE")
    return output_path


def image_only_scan_pdf_factory(
    output_path: Path,
    *,
    form_id: str = "W-2",
) -> Path:
    """Generate an image-only PDF (no text layer) — OCR path fixture.

    Renders labeled boxes to a PNG via Pillow, then embeds that PNG as
    the only content of a PDF page. Task 9 exercises this through the
    Textract OCR adapter.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Create PNG with form-like content
    png_path = output_path.with_suffix(".tmp.png")
    img = Image.new("RGB", (1224, 1584), "white")  # 8.5x11 @ 144 DPI
    draw = ImageDraw.Draw(img)

    draw.rectangle([(50, 50), (1174, 130)], outline="black", width=3)
    draw.text((60, 75), f"Synthetic {form_id} - Image-Only Scan", fill="black")
    draw.text((60, 100), "No embedded text layer. OCR required.", fill="black")

    # Fake boxes with values — just painted pixels, no text layer
    box_lines = [
        ("Box 1 Wages", "12,345.67"),
        ("Box 2 Federal withheld", "1,234.56"),
        ("Box 3 SS wages", "12,345.67"),
        ("Box 4 SS tax", "765.43"),
        ("Box 5 Medicare wages", "12,345.67"),
        ("Box 6 Medicare tax", "179.01"),
    ]
    y = 200
    for label, value in box_lines:
        draw.rectangle([(60, y), (1160, y + 110)], outline="black", width=2)
        draw.text((80, y + 20), label, fill="black")
        draw.text((900, y + 60), value, fill="black")
        y += 130

    img.save(png_path, "PNG", dpi=(144, 144))

    # Embed PNG into a PDF page
    c = rl_canvas.Canvas(str(output_path), pagesize=LETTER)
    c.drawImage(str(png_path), 0, 0, LETTER[0], LETTER[1])
    c.showPage()
    c.save()

    png_path.unlink(missing_ok=True)
    return output_path
