"""Xero Trial Balance XLSX factory.

Xero exports are XLSX with a "Trial Balance" sheet, company name + date range
in the header, and columns: Account, Code, Debit, Credit (plus a YTD column
in practice — omitted here for MVP).

Sheet title "Trial Balance" and app.xml fingerprint drive source detection.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from factories._data import DEFAULT_CHART, Account, balanced_debits_credits


def xero_tb_xlsx_factory(
    output_path: Path,
    *,
    company_name: str = "Synthetic Demo Co, LLC",
    period_end: str = "31 Dec 2024",
    accounts: Sequence[Account] | None = None,
) -> Path:
    """Generate a Xero-style Trial Balance XLSX."""
    accs = tuple(accounts) if accounts is not None else DEFAULT_CHART
    debits, credits = balanced_debits_credits(accs)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Trial Balance"

    bold = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill("solid", fgColor="E7E6E6")
    thin = Side(border_style="thin", color="808080")
    border = Border(top=thin, bottom=thin)

    # Company header
    ws["A1"] = company_name
    ws["A1"].font = Font(name="Calibri", size=14, bold=True)
    ws.merge_cells("A1:D1")

    ws["A2"] = "Trial Balance"
    ws["A2"].font = bold
    ws.merge_cells("A2:D2")

    ws["A3"] = f"As at {period_end}"
    ws.merge_cells("A3:D3")

    # Column headers on row 5
    headers = ["Account", "Code", "Debit", "Credit"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=i, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    row = 6
    for a in accs:
        ws.cell(row=row, column=1, value=a.name)
        ws.cell(row=row, column=2, value=a.number)
        ws.cell(row=row, column=3, value=float(a.balance) if a.normal_balance == "debit" else None)
        ws.cell(row=row, column=4, value=float(a.balance) if a.normal_balance == "credit" else None)
        ws.cell(row=row, column=3).number_format = "#,##0.00"
        ws.cell(row=row, column=4).number_format = "#,##0.00"
        row += 1

    # Total row
    total_row = row
    ws.cell(row=total_row, column=1, value="Total").font = bold
    ws.cell(row=total_row, column=3, value=float(debits)).font = bold
    ws.cell(row=total_row, column=4, value=float(credits)).font = bold
    ws.cell(row=total_row, column=3).number_format = "#,##0.00"
    ws.cell(row=total_row, column=4).number_format = "#,##0.00"
    ws.cell(row=total_row, column=3).border = Border(top=Side(border_style="medium"))
    ws.cell(row=total_row, column=4).border = Border(top=Side(border_style="medium"))

    # Column widths
    widths = {1: 40, 2: 12, 3: 16, 4: 16}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Xero fingerprint — creator string
    wb.properties.creator = "Xero"
    wb.properties.title = f"Trial Balance - {company_name}"
    wb.properties.subject = "Trial Balance"

    wb.save(output_path)
    return output_path
