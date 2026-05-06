"""CCH Axcess Engagement import template XLSX factory.

Produces the 10-column workpaper layout CCH Engagement accepts for import
(empty template, no balances filled in). Columns mirror the documented
CCH Engagement TB Structure per design §R17:

    Account Number | Account Name | Account Type | Prior Year |
    Unadjusted | AJE | Adjusted | RJE | Final | TJE | Tax Basis |
    Financial Statement Grouping | Tax Grouping

This is our synthetic approximation, not redistributed vendor content.
See tests/fixtures/vendor/README.md for licensing rationale.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


CCH_ENGAGEMENT_COLUMNS = (
    "Account Number",
    "Account Name",
    "Account Type",
    "Prior Year",
    "Unadjusted",
    "AJE",
    "Adjusted",
    "RJE",
    "Final",
    "TJE",
    "Tax Basis",
    "Financial Statement Grouping",
    "Tax Grouping",
)


def cch_engagement_import_xlsx_factory(
    output_path: Path,
    *,
    company_name: str = "Synthetic Demo Co, LLC",
    engagement_year: int = 2024,
) -> Path:
    """Generate an empty CCH Axcess Engagement import template."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "TB Import"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="B4C7E7")
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = f"{company_name} - {engagement_year} Trial Balance Import"
    ws["A1"].font = Font(size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(CCH_ENGAGEMENT_COLUMNS))

    ws["A2"] = (
        "CCH Axcess Engagement Import Template (synthetic approximation). "
        "Do not edit column headers. Populate rows starting on row 5."
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(CCH_ENGAGEMENT_COLUMNS))

    # Header row
    for i, col in enumerate(CCH_ENGAGEMENT_COLUMNS, start=1):
        cell = ws.cell(row=4, column=i, value=col)
        cell.font = bold
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Column widths — wide for account name, narrower for amounts
    widths: dict[int, int] = {
        1: 15, 2: 32, 3: 14, 4: 14, 5: 14, 6: 12, 7: 14, 8: 12, 9: 14,
        10: 12, 11: 14, 12: 26, 13: 26,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[4].height = 30

    wb.properties.creator = "accounting-parser"
    wb.properties.title = f"CCH Engagement Import Template - {company_name} {engagement_year}"

    wb.save(output_path)
    return output_path
