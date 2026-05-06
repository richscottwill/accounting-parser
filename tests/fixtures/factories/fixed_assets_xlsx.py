"""Fixed-asset schedule XLSX factory (BNA-style).

Generates a depreciation schedule with one row per asset across the columns
Task 16 (Depreciation Engine) parses: description, class life, placed-in-
service date, cost basis, Section 179 election, bonus depreciation, prior
accumulated, current-year book dep, current-year tax dep.
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def fixed_asset_schedule_factory(
    output_path: Path,
    *,
    company_name: str = "Synthetic Demo Co, LLC",
    tax_year: int = 2024,
) -> Path:
    """Generate a fixed-asset depreciation schedule XLSX."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Fixed Assets"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9D9D9")

    ws["A1"] = f"{company_name} - Fixed Asset Schedule - {tax_year}"
    ws["A1"].font = Font(size=13, bold=True)
    ws.merge_cells("A1:L1")

    headers = (
        "Asset ID", "Description", "Class Life (yrs)", "Placed in Service",
        "Cost Basis", "Section 179", "Bonus %", "Bonus Amount",
        "Prior Accum Dep", "Current Year Book Dep", "Current Year Tax Dep",
        "Ending Accum Dep",
    )
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = bold
        c.fill = header_fill

    # Deterministic sample assets: pre- and post- OBBBA Jan 20, 2025 split
    sample_assets = [
        ("FA001", "Office Desk",                    7,  date(2023, 6, 15),  Decimal("2500.00"),   Decimal("0"),     "40",  Decimal("1000.00"),  Decimal("214.29"),   Decimal("357.14"),   Decimal("178.57"),   None),
        ("FA002", "Delivery Van",                   5,  date(2024, 3, 10),  Decimal("35000.00"),  Decimal("0"),     "60",  Decimal("21000.00"), Decimal("0"),        Decimal("7000.00"),  Decimal("2800.00"),  None),
        ("FA003", "Server Rack Equipment",          5,  date(2024, 11, 1),  Decimal("12500.00"),  Decimal("12500"), "0",   Decimal("0"),        Decimal("0"),        Decimal("0"),        Decimal("0"),        None),
        ("FA004", "Building Improvement",           39, date(2022, 4, 30),  Decimal("85000.00"),  Decimal("0"),     "0",   Decimal("0"),        Decimal("5558.97"),  Decimal("2179.49"),  Decimal("2179.49"),  None),
        ("FA005", "Warehouse Forklift",             7,  date(2025, 1, 15),  Decimal("28000.00"),  Decimal("0"),     "40",  Decimal("11200.00"), Decimal("0"),        Decimal("4000.00"),  Decimal("2400.00"),  None),  # pre-OBBBA
        ("FA006", "Workstations (5)",               5,  date(2025, 3, 1),   Decimal("15000.00"),  Decimal("0"),     "100", Decimal("15000.00"), Decimal("0"),        Decimal("3000.00"),  Decimal("0"),        None),  # post-OBBBA
    ]

    row = 4
    for asset in sample_assets:
        for col, value in enumerate(asset, start=1):
            # Col 12 (ending accum dep) = prior + current tax
            if col == 12:
                prior = asset[8] if asset[8] is not None else Decimal("0")
                curr = asset[10] if asset[10] is not None else Decimal("0")
                value = prior + curr
            cell = ws.cell(row=row, column=col, value=value if not isinstance(value, Decimal) else float(value))
            if col in (5, 6, 8, 9, 10, 11, 12):
                cell.number_format = "#,##0.00"
            if col == 4 and isinstance(value, date):
                cell.number_format = "mm/dd/yyyy"
        row += 1

    widths = {1: 10, 2: 30, 3: 10, 4: 14, 5: 14, 6: 12, 7: 8, 8: 14, 9: 16, 10: 18, 11: 18, 12: 14}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    wb.properties.creator = "BNA Fixed Assets (synthetic)"
    wb.properties.title = f"Fixed Asset Schedule - {tax_year}"

    wb.save(output_path)
    return output_path
